# -*- coding: utf-8 -*-
"""T01 刮削准备度 Excel 检查表生成脚本

读取 ImportPlan JSON，生成 scrape_review.xlsx。
用于 M08 刮削前检查数据质量。

用法:
    python scripts/tools/export_scrape_review_excel.py
    python scripts/tools/export_scrape_review_excel.py --plan path/to/plan.json --output data/reports/scrape_review.xlsx
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

# 确保可以导入 app 模块
sys.path.insert(0, str(Path(__file__).parent.parent / "backend"))


# ============================================================
# 常量
# ============================================================

# 字幕组/技术标签关键词
_FANSUB_TECH_KEYWORDS = [
    "bdrip", "web-dl", "webrip", "1080p", "2160p", "720p",
    "hevc", "h264", "h265", "x264", "x265", "avc",
    "flac", "aac", "ac3", "truehd", "dts",
    "chs", "cht", "jpn", "eng", "chi_jpn",
    "mp4", "mkv", "ma10p", "vcb", "raws", "sub",
    "字幕组", "10bit", "8bit",
]

# group_type 排序权重
_GROUP_TYPE_ORDER = {"season": 0, "sps": 1, "op_ed": 2, "movie": 3, "": 4}
_NON_SCRAPE_DIRS = {"OP＆ED", "OP&ED", "OPED", "OP_ED", "SPs", "SPS", "SP", "Special", "special"}

# openpyxl 是导出 Excel 时才需要的可选依赖；纯规则测试不应强制安装。
Workbook = None
Alignment = None
get_column_letter = None
_HEADER_FILL = None
_HEADER_FONT_WHITE = None
_RISK_FILL_RED = None
_RISK_FILL_YELLOW = None
_GROUP_FILL_EVEN = None
_GROUP_FILL_ODD = None
_THIN_BORDER = None


def _ensure_openpyxl() -> None:
    """Load openpyxl lazily so parsing helpers remain importable without it."""
    global Workbook, Alignment, get_column_letter
    global _HEADER_FILL, _HEADER_FONT_WHITE, _RISK_FILL_RED, _RISK_FILL_YELLOW
    global _GROUP_FILL_EVEN, _GROUP_FILL_ODD, _THIN_BORDER
    if Workbook is not None:
        return
    try:
        from openpyxl import Workbook as _Workbook
        from openpyxl.styles import Alignment as _Alignment
        from openpyxl.styles import Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter as _get_column_letter
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError("导出 Excel 需要安装 openpyxl：pip install openpyxl") from exc

    Workbook = _Workbook
    Alignment = _Alignment
    get_column_letter = _get_column_letter
    _HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
    _RISK_FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    _RISK_FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    _GROUP_FILL_EVEN = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    _GROUP_FILL_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    _THIN_BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )


# ============================================================
# 数据加载
# ============================================================

def find_latest_plan() -> Path:
    """查找最新的 ImportPlan JSON"""
    plans_dir = Path(__file__).parent.parent / "backend" / "data" / "import_plans"
    if not plans_dir.exists():
        raise FileNotFoundError(f"import_plans 目录不存在: {plans_dir}")

    # 优先找 executed，再找 confirmed
    for status in ("executed", "confirmed"):
        for f in sorted(plans_dir.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True):
            if "_latest" in f.name:
                continue
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                if data.get("status") == status:
                    return f
            except (json.JSONDecodeError, KeyError):
                continue

    # 任意最新
    json_files = [f for f in plans_dir.glob("*.json") if "_latest" not in f.name]
    if json_files:
        return max(json_files, key=lambda p: p.stat().st_mtime)

    raise FileNotFoundError("未找到 ImportPlan JSON 文件")


def load_plan(path: Path) -> dict:
    """加载 ImportPlan JSON"""
    return json.loads(path.read_text(encoding="utf-8"))


# ============================================================
# 刮削启用 / 显示名 / 刮削候选
# ============================================================

def _is_scrape_enabled(item: dict) -> bool:
    """判断条目是否启用刮削

    season -> 是
    movie  -> 是
    sps    -> 否
    op_ed  -> 否
    """
    gt = item.get("group_type", "")
    return gt in ("season", "movie")


def _build_display_title(item: dict) -> str:
    """构建本地显示名

    OP/ED: NCOP01 / NCED03 / OP / ED
    SPs: 从 title/subwork_dir/filename 提取，清理技术标签
    season/movie: 不需要单独显示名（用 work_title）
    """
    gt = item.get("group_type", "")
    title = item.get("title", "")

    if gt == "op_ed":
        # OP/ED 显示名：优先用 title，否则从文件名提取
        if title:
            return title
        # 从文件名提取 OP/ED 标记
        fn = item.get("relative_path", "").split("/")[-1]
        m = re.search(r"(NCOP\d*|NCED\d*|OP\d*|ED\d*)", fn, re.IGNORECASE)
        if m:
            return m.group(1).upper()
        return ""

    if gt == "sps":
        # SPs 显示名：清理技术标签
        display = title
        if not display:
            display = _extract_subwork_dir(item)
        if not display:
            # 从文件名提取
            fn = item.get("relative_path", "").split("/")[-1]
            display = fn
        # 清理技术标签
        display = _clean_display_title(display)
        return display

    # season/movie: 不需要单独显示名
    return ""


def _clean_display_title(text: str) -> str:
    """清理显示名中的技术标签"""
    if not text:
        return ""
    cleaned = text
    # 去掉 [字幕组名] 标签
    cleaned = re.sub(r"\[[^\]]*(?:Sub|Raw|Raws|Studio|VCB|LoliHouse|Sakurato|DMG|Haruhana|Nekomoe|kissaten|BeanSub|FZSD|MAI|CZ|NW|EA|AS|LP|Sweet|THX|T\.H\.X)[^\]]*\]", "", cleaned, flags=re.IGNORECASE)
    # 去掉 [技术参数] 标签
    cleaned = re.sub(r"\[(?:BDRip|WebRip|WEB-DL|Ma\d+[pi]_\d+[pi]|x\d+_\w+|HEVC[^\]]*|H\d+[^\]]*|FLAC|AAC|AC3|CHS|CHT|JPN|CHI_JPN|1080P|2160P|720P|MP4|MKV|Fin|简繁[^\]]*)[^\]]*\]", "", cleaned, flags=re.IGNORECASE)
    # 去掉 (技术参数)
    cleaned = re.sub(r"\((?:BDRip|WebRip|1080p|2160p|HEVC[^\)]*|FLAC|AAC|AC3)[^\)]*\)", "", cleaned, flags=re.IGNORECASE)
    # 去掉编号前缀（如 5.）
    cleaned = re.sub(r"^\d+\.", "", cleaned)
    # 去掉年份
    cleaned = re.sub(r"[.．]\d{4}$", "", cleaned)
    cleaned = re.sub(r"[(\（]\d{4}[)\）]$", "", cleaned)
    # 清理多余空格和标点
    cleaned = " ".join(cleaned.split())
    cleaned = cleaned.strip(" .-_")
    return cleaned


def _iter_notes(item: dict):
    """遍历 reasons/warnings。子作品目录已降噪到 reasons，但旧数据可能仍在 warnings。"""
    for key in ("reasons", "warnings"):
        for note in item.get(key, []) or []:
            yield str(note)


def _extract_subwork_dir(item: dict) -> str:
    """提取子作品目录线索，供 T01R 和 M08 前置检查使用。"""
    for note in _iter_notes(item):
        if "子作品目录:" in note:
            candidate = note.split("子作品目录:", 1)[1].strip()
            if candidate and candidate not in _NON_SCRAPE_DIRS:
                return candidate

    # 兼容历史计划：如果 reasons/warnings 没有线索，则从真实路径回推父目录。
    # 只在父目录和 original_title 不同的时候采纳，避免把作品容器误当子作品。
    original_title = item.get("original_title", "")
    path_text = item.get("real_path") or item.get("relative_path") or ""
    parts = [p for p in re.split(r"[\\/]+", path_text) if p]
    if len(parts) >= 2:
        parent = parts[-2]
        if parent and parent != original_title and parent not in _NON_SCRAPE_DIRS:
            return parent
    return ""


def _extract_year_from_text(text: str):
    """从目录名中提取年份。"""
    if not text:
        return None
    for pat in (r"[.．](\d{4})(?:$|[^\d])", r"[(（](\d{4})[)）]"):
        m = re.search(pat, text)
        if m:
            year = int(m.group(1))
            if 1900 <= year <= 2099:
                return year
    return None


def _clean_scrape_title(text: str, group_type: str = "") -> str:
    """清理刮削搜索标题。

    候选名只负责搜索标题，不承担季号区分；季号保留在 season_number。
    """
    if not text:
        return ""

    cleaned = _clean_display_title(text)

    # 去掉季号标记，但保留官方副标题，如 CLANNAD After Story、超电磁炮T。
    cleaned = re.sub(r"[.．_ -]*[\[【(（][Ss]\d{1,2}[\]】)）]", " ", cleaned)
    cleaned = re.sub(r"[.．_ -]+[Ss]\d{1,2}(?=$|[.．_ -])", " ", cleaned)
    cleaned = re.sub(r"\s+[Ss]\d{1,2}(?=$|\s)", " ", cleaned)
    cleaned = re.sub(r"\bSeason\s*\d{1,2}\b", " ", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"第\s*\d+\s*季", " ", cleaned)

    # 去掉结尾文件扩展名残留。
    cleaned = re.sub(r"\.(mkv|mp4|avi|mov|wmv|flv|m2ts|ts)$", "", cleaned, flags=re.IGNORECASE)

    cleaned = re.sub(r"\s+", " ", cleaned)
    cleaned = cleaned.strip(" .-_")
    return cleaned


def _extract_scrape_candidate(item: dict) -> tuple:
    """提取刮削候选名和年份（仅 season/movie）

    返回: (candidate_name, candidate_year)
    """
    # OP/ED 和 SPs 不生成刮削候选
    gt = item.get("group_type", "")
    if gt in ("op_ed", "sps"):
        return "", None

    subwork_dir = _extract_subwork_dir(item)

    # 候选名：优先从子作品目录清洗，否则用 work_title
    candidate_name = _clean_scrape_title(item.get("work_title", ""), gt)
    if subwork_dir:
        cleaned = _clean_scrape_title(subwork_dir, gt)
        if cleaned:
            candidate_name = cleaned

    # 候选年份：优先 item.year，否则从子作品目录提取
    candidate_year = item.get("year")
    if candidate_year is None and subwork_dir:
        candidate_year = _extract_year_from_text(subwork_dir)

    return candidate_name, candidate_year


def _extract_non_scrape_reason(item: dict) -> str:
    """提取非刮削原因"""
    gt = item.get("group_type", "")
    if gt == "op_ed":
        return "OP/ED 不生成作品级刮削目标"
    if gt == "sps":
        return "SPs 默认不生成作品级刮削目标"
    return ""


def _detect_fansub_tech(text: str) -> bool:
    """检测文本是否含字幕组/技术标签"""
    if not text:
        return False
    lower = text.lower()
    for kw in _FANSUB_TECH_KEYWORDS:
        if kw in lower:
            return True
    return False


# ============================================================
# 风险标记
# ============================================================

def _assess_risks(item: dict, series_years: dict) -> list:
    """评估单个 item 的风险（只报真正需要关注的问题）"""
    risks = []
    gt = item.get("group_type", "")

    # OP/ED 和 SPs：只检查 display_title
    if gt in ("op_ed", "sps"):
        display = _build_display_title(item)
        if not display:
            risks.append(f"{gt} 缺显示名")
        if item.get("needs_review"):
            risks.append("needs_review")
        return risks

    # season/movie：检查真正的问题
    # 重复集号 warning
    for w in item.get("warnings", []):
        if "多个来源文件" in w:
            risks.append(w)
            break

    # needs_review
    if item.get("needs_review"):
        risks.append("needs_review")

    # series_group 含脏标签
    sg = item.get("series_group", "")
    if _detect_fansub_tech(sg):
        risks.append("series_group 含脏标签")

    # work_title 含脏标签
    wt = item.get("work_title", "")
    if _detect_fansub_tech(wt):
        risks.append("work_title 含脏标签")

    # main_series 但 series_group 为空
    if item.get("card_type") == "main_series" and not sg:
        risks.append("main_series 缺 series_group")

    # season 缺必要字段
    if gt == "season":
        if item.get("season_number") is None:
            risks.append("season 缺 season_number")
        if item.get("episode_number") is None:
            risks.append("season 缺 episode_number")

    return risks


# ============================================================
# Excel 生成
# ============================================================

def _apply_header_style(ws, col_count):
    """应用表头样式"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT_WHITE
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER


def _auto_width(ws, max_width=28):
    """自动列宽"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def generate_excel(plan: dict, output_path: Path):
    """生成 Excel"""
    _ensure_openpyxl()
    wb = Workbook()
    items = plan.get("items", [])

    # 只处理 video + generate_strm
    video_items = [
        it for it in items
        if it.get("resource_type") == "video" and it.get("action") == "generate_strm"
    ]

    # 预计算每个 series_group 的年份集合
    series_years = defaultdict(set)
    for it in video_items:
        sg = it.get("series_group", "")
        yr = it.get("year")
        if sg and yr:
            series_years[sg].add(yr)

    # ============================================================
    # Sheet 1: Overview
    # ============================================================
    ws_overview = wb.active
    ws_overview.title = "Overview"

    main_series_count = len(set(it.get("series_group", "") for it in video_items if it.get("card_type") == "main_series" and it.get("series_group")))
    standalone_count = sum(1 for it in video_items if it.get("card_type") == "standalone")
    missing_year = sum(1 for it in video_items if it.get("year") is None)
    has_warning = sum(1 for it in video_items if it.get("warnings"))
    risk_count = sum(1 for it in video_items if _assess_risks(it, series_years))

    overview_data = [
        ["字段", "值"],
        ["plan_id", plan.get("plan_id", "")],
        ["source", plan.get("source", "")],
        ["plan_status", plan.get("status", "")],
        ["视频条目数", len(video_items)],
        ["主系列数量", main_series_count],
        ["独立卡片数量", standalone_count],
        ["缺年份数量", missing_year],
        ["有 warning 数量", has_warning],
        ["疑似刮削风险数量", risk_count],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    for row in overview_data:
        ws_overview.append(row)
    _apply_header_style(ws_overview, 2)
    _auto_width(ws_overview)

    # ============================================================
    # Sheet 2: Scrape Review
    # ============================================================
    ws_review = wb.create_sheet("Scrape Review")

    headers = [
        "系列", "类型", "作品名", "年份", "分组", "季", "集",
        "显示名", "可刮削", "候选名", "候选年", "非刮削原因",
        "原始标题", "子目录", "真实路径", "置信度", "需确认",
        "警告", "风险", "备注",
    ]
    ws_review.append(headers)
    _apply_header_style(ws_review, len(headers))

    # 排序
    def sort_key(it):
        sg = it.get("series_group", "") or it.get("work_title", "")
        ct = it.get("card_type", "")
        gt = it.get("group_type", "")
        sn = it.get("season_number") or 999
        en = it.get("episode_number") or 999
        sp = it.get("special_number") or 999
        rp = it.get("relative_path", "")
        return (sg, ct, _GROUP_TYPE_ORDER.get(gt, 9), sn, en, sp, rp)

    sorted_items = sorted(video_items, key=sort_key)

    prev_series = None
    row_idx = 1  # 从 1 开始（表头已占 1）
    group_color_toggle = False

    for it in sorted_items:
        sg = it.get("series_group", "") or it.get("work_title", "")

        # 不同系列之间插入空行
        if prev_series is not None and sg != prev_series:
            ws_review.append([])
            row_idx += 1
            group_color_toggle = not group_color_toggle
        prev_series = sg

        subwork_dir = _extract_subwork_dir(it)

        candidate_name, candidate_year = _extract_scrape_candidate(it)
        display_title = _build_display_title(it)
        scrape_enabled = _is_scrape_enabled(it)
        non_scrape_reason = _extract_non_scrape_reason(it)
        risks = _assess_risks(it, series_years)

        row = [
            sg,
            it.get("card_type", ""),
            it.get("work_title", ""),
            it.get("year"),
            it.get("group_type", ""),
            it.get("season_number"),
            it.get("episode_number"),
            display_title,
            "是" if scrape_enabled else "否",
            candidate_name if scrape_enabled else "",
            candidate_year if scrape_enabled else "",
            non_scrape_reason,
            it.get("original_title", ""),
            subwork_dir,
            it.get("real_path", ""),
            it.get("confidence", ""),
            "✓" if it.get("needs_review") else "",
            "; ".join(it.get("warnings", [])),
            "; ".join(risks),
            "",
        ]
        ws_review.append(row)
        row_idx += 1

        # 风险行着色
        if risks:
            fill = _RISK_FILL_YELLOW if "needs_review" in "; ".join(risks) else _RISK_FILL_RED
            for col in range(1, len(headers) + 1):
                ws_review.cell(row=row_idx, column=col).fill = fill
        elif group_color_toggle:
            for col in range(1, len(headers) + 1):
                ws_review.cell(row=row_idx, column=col).fill = _GROUP_FILL_EVEN

        # 边框
        for col in range(1, len(headers) + 1):
            ws_review.cell(row=row_idx, column=col).border = _THIN_BORDER

    # 冻结首行，开启筛选
    ws_review.freeze_panes = "A2"
    ws_review.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx}"
    _auto_width(ws_review)

    # ============================================================
    # Sheet 3: Series Summary
    # ============================================================
    ws_series = wb.create_sheet("Series Summary")

    series_headers = [
        "系列", "目录", "视频", "季", "SP", "OP-ED", "独立",
        "年份", "可刮削候选", "SPs 样本", "OP-ED 样本", "风险",
    ]
    ws_series.append(series_headers)
    _apply_header_style(ws_series, len(series_headers))

    # 按 series_group 聚合
    series_map = defaultdict(lambda: {
        "items": [], "seasons": set(), "sps": 0, "op_ed": 0, "standalone": 0,
        "years": set(), "scrape_candidates": set(), "sps_displays": set(), "op_ed_displays": set(),
    })

    for it in video_items:
        sg = it.get("series_group", "") or it.get("work_title", "")
        entry = series_map[sg]
        entry["items"].append(it)
        gt = it.get("group_type", "")
        if gt == "season":
            sn = it.get("season_number")
            if sn is not None:
                entry["seasons"].add(sn)
        elif gt == "sps":
            entry["sps"] += 1
            display = _build_display_title(it)
            if display:
                entry["sps_displays"].add(display[:30])
        elif gt == "op_ed":
            entry["op_ed"] += 1
            display = _build_display_title(it)
            if display:
                entry["op_ed_displays"].add(display[:20])
        if it.get("card_type") == "standalone":
            entry["standalone"] += 1
        yr = it.get("year")
        if yr:
            entry["years"].add(yr)
        # 只收集可刮削条目的候选名
        if _is_scrape_enabled(it):
            cn, _ = _extract_scrape_candidate(it)
            if cn:
                entry["scrape_candidates"].add(cn[:30])

    for sg, entry in sorted(series_map.items()):
        # 风险标记
        series_risks = []
        sg_years = series_years.get(sg, set())
        if len(sg_years) > 1:
            series_risks.append(f"含 {len(sg_years)} 个年份")
        if _detect_fansub_tech(sg):
            series_risks.append("series_group 含脏标签")

        # mirror 目录
        mirror_dir = sg
        for it in entry["items"]:
            tsp = it.get("target_strm_path", "")
            if tsp:
                parts = Path(tsp).parts
                for i, p in enumerate(parts):
                    if p == "115" and i + 1 < len(parts):
                        mirror_dir = parts[i + 1]
                        break
                break

        row = [
            sg,
            mirror_dir,
            len(entry["items"]),
            len(entry["seasons"]),
            entry["sps"],
            entry["op_ed"],
            entry["standalone"],
            ", ".join(str(y) for y in sorted(entry["years"])),
            "; ".join(list(entry["scrape_candidates"])[:3]),
            "; ".join(list(entry["sps_displays"])[:3]),
            "; ".join(list(entry["op_ed_displays"])[:3]),
            "; ".join(series_risks),
        ]
        ws_series.append(row)

    _auto_width(ws_series)

    # ============================================================
    # Sheet 4: Issues
    # ============================================================
    ws_issues = wb.create_sheet("Issues")

    issue_headers = [
        "系列", "作品名", "分组", "季", "集",
        "风险", "警告", "需确认", "候选名", "候选年",
    ]
    ws_issues.append(issue_headers)
    _apply_header_style(ws_issues, len(issue_headers))

    issue_items = []
    for it in video_items:
        gt = it.get("group_type", "")
        risks = _assess_risks(it, series_years)

        # 降噪：只保留真正需要用户关注的问题
        gt = it.get("group_type", "")
        real_issues = []

        # OP/ED 和 SPs：只有 needs_review、缺显示名才进
        if gt in ("op_ed", "sps"):
            display = _build_display_title(it)
            if it.get("needs_review"):
                real_issues.append("needs_review")
            if not display:
                real_issues.append(f"{gt} 缺显示名")
            if not real_issues:
                continue
            risks = real_issues
        else:
            # season/movie：只保留重复集号、缺显示名、needs_review、识别冲突
            for r in risks:
                if any(kw in r for kw in ["多个来源文件", "缺显示名", "needs_review", "冲突"]):
                    real_issues.append(r)
            if not real_issues:
                continue
            risks = real_issues

        if risks:
            candidate_name, candidate_year = _extract_scrape_candidate(it)
            sg = it.get("series_group", "") or it.get("work_title", "")
            issue_items.append([
                sg,
                it.get("work_title", ""),
                gt,
                it.get("season_number"),
                it.get("episode_number"),
                "; ".join(risks),
                "; ".join(it.get("warnings", [])),
                "✓" if it.get("needs_review") else "",
                candidate_name,
                candidate_year,
            ])

    # 按风险数量排序（多的在前）
    issue_items.sort(key=lambda r: r[0])

    for row_data in issue_items:
        ws_issues.append(row_data)

    _auto_width(ws_issues)

    # ============================================================
    # 保存
    # ============================================================
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


# ============================================================
# 主入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="生成刮削准备度 Excel 检查表")
    parser.add_argument("--plan", type=str, help="ImportPlan JSON 路径")
    parser.add_argument("--output", type=str, default="data/reports/scrape_review.xlsx", help="输出路径")
    args = parser.parse_args()

    # 加载 plan
    if args.plan:
        plan_path = Path(args.plan)
    else:
        plan_path = find_latest_plan()

    print(f"读取 ImportPlan: {plan_path}")
    plan = load_plan(plan_path)

    # 生成 Excel
    output_path = Path(args.output)
    generate_excel(plan, output_path)

    # 统计
    items = plan.get("items", [])
    video_items = [it for it in items if it.get("resource_type") == "video" and it.get("action") == "generate_strm"]
    series_groups = set(it.get("series_group", "") or it.get("work_title", "") for it in video_items)

    print(f"输出: {output_path}")
    print(f"视频条目: {len(video_items)}")
    print(f"系列/作品数: {len(series_groups)}")
    print(f"Excel 工作表: Overview / Scrape Review / Series Summary / Issues")


if __name__ == "__main__":
    main()
